# -*- coding: utf-8 -*-
from __future__ import unicode_literals

import datetime
import uuid
from collections import OrderedDict
import openpyxl
import os
import json

from docx import Document

from django.contrib.gis.geos import Point
from django.conf import settings
from django.contrib.auth import authenticate, login, logout
from django.core.mail import send_mail
from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt


from rest_framework.views import APIView
from rest_framework import status, parsers, renderers
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework.authtoken.views import ObtainAuthToken, Token
from rest_framework.authtoken.serializers import AuthTokenSerializer
from rest_framework.permissions import AllowAny

from .serializers import SignupSerializer, CategorySerializer
from .models import User, UserProfile
from catalogs.models import Category, ExpirationValue

from customer.models import ProductCard
from .models import UserProfile, Address, OrganizationType, LegalAct, SignerInfo, IdentityDocument, Account
from payments.models import  DirectPayment
from customer.models import TradePoint, OrderUnit, Order
from producer.models import ProducerDepot
from payments.models import OrderPayment, PaymentNotification, Success

@api_view(['POST'])
def sign_up(request):
    serialized_user = SignupSerializer(data=request.data)
    if serialized_user.is_valid():
        registered_user = User.objects.create_user(
            email=request.data['email'],
            username=request.data['email'],
            password=request.data['password']
        )
        UserProfile.objects.create(
            user=registered_user,
            inn=request.data['inn'],
            phone=request.data['phone'],
            role=request.data['role'],
            company_name=request.data['company_name']
        )
        user = authenticate(username=request.data['email'], password=request.data['password'])
        if user is not None:
            login(request, user)

        email_message = u"Зарегистрирован новый пользователь\n"
        email_message += u"Название компании: " + request.data['company_name'] + "\n"
        email_message += u"email: " + request.data['email'] + "\n"
        email_message += u"ИНН: " + request.data['inn'] + "\n"
        email_message += u"Телефон: " + request.data['phone'] + "\n"

        if serialized_user.validated_data['role'] == 'producer':
            email_message += u"Роль: поставщик\n"
        elif serialized_user.validated_data['role'] == 'customer':
            email_message += u"Роль: торговая точка\n"

        if request.get_host() != '127.0.0.1:8000':
            send_mail(u'Регистрация нового пользователя на сайте the-sklad.ru', email_message, settings.EMAIL_HOST_USER, ['ceo@the-sklad.ru', ])
        return Response(serialized_user.data, status=status.HTTP_201_CREATED)
    else:
        return Response(serialized_user.errors, status=status.HTTP_400_BAD_REQUEST)


class SignIn(APIView):
    throttle_classes = ()
    permission_classes = (AllowAny,)
    parser_classes = (parsers.FormParser, parsers.MultiPartParser, parsers.JSONParser,)
    renderer_classes = (renderers.JSONRenderer,)
    serializer_class = AuthTokenSerializer

    def post(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data['user']
        token, created = Token.objects.get_or_create(user=user)
        user = authenticate(username=request.data['username'], password=request.data['password'])
        if user is not None:
            login(request, user)
        store = {'token': token.key}
        if hasattr(request.user, 'profile'):
            store['role'] = request.user.profile.role
        return Response(store)


def logout(request):
    logout(request)
    return JsonResponse({'role': None, 'token': None})


def categories(request, pk=None):
    final_data = {'categories': []}
    for cat in Category.objects.filter(pid__isnull=True, disabled=False) if pk is None else Category.objects.filter(pid=pk, disabled=False):
        final_unseen = 0
        for sub_cat in Category.objects.filter(pid=cat.id):
            final_unseen += sub_cat.productcard_set.count() - sub_cat.productcard_set.filter(
                seen__id=request.user.id).count()

        final_data['categories'].append({
            'id': cat.id,
            'name': cat.name,
            'image': cat.get_image_url(),
            'unseen': final_unseen
        })
        if pk:
            final_data['current_cat'] = Category.objects.get(pk=pk).name
    return JsonResponse(final_data)

def index(request):
    if request.user.is_anonymous:
        if request.method == 'GET':
            return render(request, 'sign_in.html')                                             #вопрос
    else:
        if hasattr(request.user, 'profile'):
            if request.user.profile.role == 'customer':
                return redirect('/categories')
            if request.user.profile.role == 'producer':
                return redirect('/my_products')
            if request.user.profile.role == 'supervisor':
                return redirect('/my_trade_points')
            if request.user.profile.role == 'manager':
                if request.user.profile.created:
                    return redirect(my_clients)
                return redirect(my_personal_data)
    return redirect('/categories')


def register(request):
    if not request.user.is_anonymous:
        return redirect(index)

    if 'role' not in request.POST:
        return render(request, 'pick_role.html')                                                    #вопрос
    role = request.POST['role']
    return JsonResponse({'role': role})


def products(request, cat_id):
    current_category = Category.objects.get(pk=cat_id)
    if hasattr(request.user, 'profile'):
        if request.user.profile.role == 'customer':
            for product in ProductCard.objects.filter(category=current_category):
                product.seen.add(request.user.id)
                product.save()

    return JsonResponse({
        'category': {
            'pid': current_category.pid_id,
            'name': current_category.name,
        },
        'products': [{
            'id': product.id,
            'name': product.name,
            'minimum_amount': product.minimum_amount,
            'description': product.description,
            'image': product.get_image_url(),
            'price': product.producer_price,
            'weight': product.weight,
        } for product in ProductCard.objects.filter(category_id=cat_id)]
    })
    # 'trade_points': TradePoint.objects.filter(customer_id=request.user.id)


@login_required(login_url='/sign_in/')
def profile(request):
    if not hasattr(request.user, 'profile'):
        return JsonResponse({'success': False, 'error_message': u'У вас нет профиля пользователя'})

    data = {
        'profile': request.user.profile,
        'organization_types': OrganizationType.objects.all(),
        'legal_acts': LegalAct.objects.all(),
        'accounts': Account.objects.filter(profile_id=request.user.profile.id),
    }
    if request.user.profile.role in ('customer', 'producer'):
        if request.user.profile.role == 'customer':
            data['trade_points'] = TradePoint.objects.filter(customer_id=request.user.id)
            #return render(request, 'profile_create_customer.html', data)
            return  JsonResponse(data)

        if request.user.profile.role == 'producer':
            data['depots'] = ProducerDepot.objects.filter(producer_id=request.user.id)
            #return render(request, 'profile_create_producer.html', data)
            return JsonResponse(data)

    if request.user.profile.role == 'manager':
        return redirect(my_personal_data)

@login_required(login_url='/sign_in')
def profile_skip_creation(request):
    if hasattr(request.user, 'profile'):
        my_profile = UserProfile.objects.get(pk=request.user.profile.id)
        my_profile.created = True
        my_profile.save()
        return redirect(profile)
    return JsonResponse({'success': False, 'error_message': u'Не найден профиль пользователя'})


@login_required(login_url='/sign_in/')
def profile_fiz_and_jur_address(request):
    u_p = request.user.profile
    if u_p:
        if u_p.juridical_address is None:
            u_p.juridical_address = Address.objects.create(
                full_address=request.POST['jur_full_address'],
                index=request.POST['jur_index'],
                region=request.POST['jur_region'],
                city=request.POST['jur_city'],
                street=request.POST['jur_street'],
                house=request.POST['jur_house'],
                block=request.POST['jur_block'],
                structure=request.POST['jur_structure'],
                flat=request.POST['jur_flat']
            )
            u_p.save()
        else:
            u_p.juridical_address.full_address = request.POST['jur_full_address']
            u_p.juridical_address.index = request.POST['jur_index']
            u_p.juridical_address.region = request.POST['jur_region']
            u_p.juridical_address.city = request.POST['jur_city']
            u_p.juridical_address.street = request.POST['jur_street']
            u_p.juridical_address.house = request.POST['jur_house']
            u_p.juridical_address.block = request.POST['jur_block']
            u_p.juridical_address.structure = request.POST['jur_structure']
            u_p.juridical_address.flat = request.POST['jur_flat']
            u_p.juridical_address.save()

        if 'physical_is_juridical' in request.POST:
            u_p.physical_is_juridical = True
        else:
            u_p.physical_is_juridical = False
            if u_p.physical_address is None:
                u_p.physical_address = Address.objects.create(
                    full_address=request.POST['fiz_full_address'],
                    index=request.POST['fiz_index'],
                    region=request.POST['fiz_region'],
                    city=request.POST['fiz_city'],
                    street=request.POST['fiz_street'],
                    house=request.POST['fiz_house'],
                    block=request.POST['fiz_block'],
                    structure=request.POST['fiz_structure'],
                    flat=request.POST['fiz_flat']
                )
                u_p.save()
            else:
                u_p.physical_address.full_address = request.POST['fiz_full_address']
                u_p.physical_address.index = request.POST['fiz_index']
                u_p.physical_address.region = request.POST['fiz_region']
                u_p.physical_address.city = request.POST['fiz_city']
                u_p.physical_address.street = request.POST['fiz_street']
                u_p.physical_address.house = request.POST['fiz_house']
                u_p.physical_address.block = request.POST['fiz_block']
                u_p.physical_address.structure = request.POST['fiz_structure']
                u_p.physical_address.flat = request.POST['fiz_flat']
                u_p.physical_address.save()
        u_p.save()

    return redirect(profile)

@login_required(login_url='/sign_in/')
def profile_juridical_address(request):
    u_p = request.user.profile
    if u_p:
        if u_p.juridical_address is None:
            u_p.juridical_address = Address.objects.create(
                full_address=request.POST['full_address'],
                index=request.POST['index'],
                region=request.POST['region'],
                city=request.POST['city'],
                street=request.POST['street'],
                house=request.POST['house'],
                block=request.POST['block'],
                structure=request.POST['structure'],
                flat=request.POST['flat']
            )
            u_p.save()
        else:
            u_p.juridical_address.full_address = request.POST['full_address']
            u_p.juridical_address.index = request.POST['index']
            u_p.juridical_address.region = request.POST['region']
            u_p.juridical_address.city = request.POST['city']
            u_p.juridical_address.street = request.POST['street']
            u_p.juridical_address.house = request.POST['house']
            u_p.juridical_address.block = request.POST['block']
            u_p.juridical_address.structure = request.POST['structure']
            u_p.juridical_address.flat = request.POST['flat']
            u_p.juridical_address.save()

        if 'physical_is_juridical' in request.POST and request.POST['physical_is_juridical'] == 'on':
            u_p.physical_is_juridical = True
        else:
            u_p.physical_is_juridical = False
        u_p.save()

        return redirect(profile)
    return JsonResponse({'success': False, 'error_message': u'Не найден профиль пользователя'})

@login_required(login_url='/sign_in/')
def profile_physical_address(request):
    u_p = request.user.profile
    if u_p:
        if u_p.physical_address is None:
            u_p.physical_address = Address.objects.create(
                index=request.POST['index'],
                region=request.POST['region'],
                city=request.POST['city'],
                street=request.POST['street'],
                house=request.POST['house'],
                block=request.POST['block'],
                structure=request.POST['structure'],
                flat=request.POST['flat']
            )
            u_p.save()
        else:
            u_p.physical_address.index = request.POST['index']
            u_p.physical_address.region = request.POST['region']
            u_p.physical_address.city = request.POST['city']
            u_p.physical_address.street = request.POST['street']
            u_p.physical_address.house = request.POST['house']
            u_p.physical_address.block = request.POST['block']
            u_p.physical_address.structure = request.POST['structure']
            u_p.physical_address.flat = request.POST['flat']
            u_p.physical_address.save()

        return redirect(profile)
    return JsonResponse({'success': False, 'error_message': u'Не найден профиль пользователя'})


@login_required(login_url='/sign_in/')
def profile_company_info(request):
    u_p = request.user.profile
    if u_p:
        u_p.inn = request.POST['inn']
        u_p.ogrn = request.POST['ogrn']
        u_p.kpp = request.POST['kpp']
        u_p.organization_type_id = request.POST['org_type_short']
        u_p.legal_act_id = request.POST['legal_act']
        u_p.save()

        return redirect(profile)
    return JsonResponse({'success': False, 'error_message': u'Не найден профиль пользователя'})


@login_required(login_url='/sign_in/')
def profile_signer_info(request):
    u_p = request.user.profile
    if u_p:
        if u_p.signer_info is None:
            u_p.signer_info = SignerInfo.objects.create(
                surname=request.POST['surname'],
                name=request.POST['name'],
                patronymic=request.POST['patronymic'],
                birth_date=datetime.datetime.strptime(request.POST['issued_date'], "%d.%m.%Y").date(),
                position=request.POST['position'],
            )
            u_p.save()
        else:
            u_p.signer_info.surname = request.POST['surname']
            u_p.signer_info.name = request.POST['name']
            u_p.signer_info.patronymic = request.POST['patronymic']
            u_p.signer_info.birth_date = datetime.datetime.strptime(request.POST['birth_date'], "%d.%m.%Y").date()
            u_p.signer_info.position = request.POST['position']
            u_p.signer_info.save()

        return redirect(profile)
    return JsonResponse({'success': False, 'error_message': u'Не найден профиль пользователя'})


@login_required(login_url='/sign_in/')
def profile_identity_document(request):
    u_p = request.user.profile
    if u_p:
        if u_p.identity_document is None:
            u_p.identity_document = IdentityDocument.objects.create(
                series=request.POST['series'],
                number=request.POST['number'],
                issued_by=request.POST['issued_by'],
                issued_date=datetime.datetime.strptime(request.POST['issued_date'], "%d.%m.%Y").date()
            )
            u_p.save()
        else:
            u_p.identity_document.series = request.POST['series']
            u_p.identity_document.number = request.POST['number']
            u_p.identity_document.issued_by = request.POST['issued_by']
            u_p.identity_document.issued_date = datetime.datetime.strptime(request.POST['issued_date'], "%d.%m.%Y").date()
            u_p.identity_document.save()
        if 'document' in request.FILES:
            # todo: При изменении файла его было бы неплохо удалять
            u_p.identity_document.document.save(str(uuid.uuid4()) + request.FILES['document'].name, request.FILES['document'])
            u_p.identity_document.save()

        return redirect(profile)
    return JsonResponse({'success': False, 'error_message': u'Не найден профиль пользователя'})

@login_required(login_url='/sign_in/')
def profile_signer_info_and_identity(request):
    u_p = request.user.profile
    if u_p:
        if u_p.signer_info is None:
            u_p.signer_info = SignerInfo.objects.create(
                surname=request.POST['surname'],
                name=request.POST['name'],
                patronymic=request.POST['patronymic'],
                birth_date=datetime.datetime.strptime(request.POST['birth_date'], "%d.%m.%Y").date(),
                position=request.POST.get('position', ''),
            )
            u_p.save()
        else:
            u_p.signer_info.surname = request.POST['surname']
            u_p.signer_info.name = request.POST['name']
            u_p.signer_info.patronymic = request.POST['patronymic']
            u_p.signer_info.birth_date = datetime.datetime.strptime(request.POST['birth_date'], "%d.%m.%Y").date()
            u_p.signer_info.position = request.POST.get('position', '')
            u_p.signer_info.save()

        if u_p.identity_document is None:
            u_p.identity_document = IdentityDocument.objects.create(
                series=request.POST['series'],
                number=request.POST['number'],
                issued_by=request.POST['issued_by'],
                issued_date=datetime.datetime.strptime(request.POST['issued_date'], "%d.%m.%Y").date()
            )
            u_p.save()
        else:
            u_p.identity_document.series = request.POST['series']
            u_p.identity_document.number = request.POST['number']
            u_p.identity_document.issued_by = request.POST['issued_by']
            u_p.identity_document.issued_date = datetime.datetime.strptime(request.POST['issued_date'], "%d.%m.%Y").date()
            u_p.identity_document.save()

        if not u_p.created:
            u_p.created = True
            u_p.save()
        if 'document' in request.FILES:
            # todo: При изменении файла его было бы неплохо удалять
            u_p.identity_document.document.save(str(uuid.uuid4()) + request.FILES['document'].name, request.FILES['document'])
            u_p.identity_document.save()

        return redirect(profile)
    return JsonResponse({'success': False, 'error_message': u'Не найден профиль пользователя'})


@login_required(login_url='/sign_in')
def profile_account_add(request):
    if hasattr(request.user, 'profile'):
        Account.objects.create(
            profile=request.user.profile,
            account_number=request.POST['account_number'],
            bik=request.POST['bik'],
            bank_name=request.POST['bank_name'],
            correspondent_account=request.POST['correspondent_account']
        )
        return redirect(profile)
    return JsonResponse({'success': False, 'error_message': u'Не найден профиль пользователя'})

@login_required(login_url='/sign_in')
def profile_account_edit(request, pk):
    if hasattr(request.user, 'profile'):
        my_account = Account.objects.get(pk=pk, profile_id=request.user.profile.id)
        if my_account:
            my_account.account_number = request.POST['account_number']
            my_account.bik = request.POST['bik']
            my_account.bank_name = request.POST['bank_name']
            my_account.correspondent_account = request.POST['correspondent_account']
            my_account.save()
            return redirect(profile)
        return JsonResponse({'success': False, 'error_message': u'Редактируемый счёт не найден'})
    return JsonResponse({'success': False, 'error_message': u'Не найден профиль пользователя'})


@login_required(login_url='/sign_in/')
def my_personal_data(request):
    if hasattr(request.user, 'profile'):
        if request.user.profile.role == 'manager':
            JsonResponse({
                'profile': request.user.profile,
                'trade_points': TradePoint.objects.filter(territory__representative__id=request.user.id)
            })
        return JsonResponse({'success': False, 'error_message': u'Только торговый представитель имеет доступ к странице'})
    return JsonResponse({'success': False, 'error_message': u'Ошибка при просмотре профиля пользователя'})


@login_required(login_url='/sign_in/')
def my_clients(request):
    if hasattr(request.user, 'profile'):
        if request.user.profile.role == 'manager':
            customers = {}
            trade_point_queryset = TradePoint.objects.filter(territory__representative__id=request.user.id)
            if 'search_string' in request.GET:
                trade_point_queryset = trade_point_queryset.filter(address__full_address__icontains=request.GET['search_string'])

            sort_type = request.GET['sort'] if 'sort' in request.GET else 'income'

            for tp in trade_point_queryset:

                if sort_type == 'name':
                    sorted_key = tp.customer.profile.company_name
                else:
                    customer_total = tp.customer.customer_total()
                    sorted_key = customer_total if customer_total != 0 else 0 - tp.customer.id

                if sorted_key in customers.keys():
                    customers[sorted_key]['trade_points'].append(tp)
                else:
                    customers[sorted_key] = {
                        "company_name": tp.customer.profile.company_name,
                        "signer_fio": (tp.customer.profile.signer_info.fio() if tp.customer.profile.signer_info else ''),
                        "phone": tp.customer.profile.nice_phone(),
                        "customer_total": tp.customer.customer_total(),
                        "trade_points": [tp, ]
                    }

                customers = OrderedDict(sorted(customers.items(), reverse=True))
            return JsonResponse({
                'profile': request.user.profile,
                'customers': customers,
                'sort_type': sort_type,
            })
        return JsonResponse({'success': False, 'error_message': u'Только торговый представитель имеет доступ к странице'})
    return JsonResponse({'success': False, 'error_message': u'Ошибка при просмотре профиля пользователя'})


@login_required(login_url='/sign_in/')
def my_income(request):
    if hasattr(request.user, 'profile'):
        if request.user.profile.role == 'manager':
            return render(request, 'manager/my_income.html')                                                #вопрос
        return JsonResponse({'success': False, 'error_message': u'Только торговый представитель имеет доступ к странице'})
    return JsonResponse({'success': False, 'error_message': u'Ошибка при просмотре профиля пользователя'})

@login_required(login_url='/sign_in/')
def my_trade_points(request):
    if hasattr(request.user, 'profile'):
        if request.user.profile.role == 'supervisor':
            final_data = {
                'gps_coordinates': [{
                    'lat': tp.address.location.y,
                    'lng': tp.address.location.x,
                    'title': tp.address.castrate_nicely()
                } for tp in TradePoint.objects.filter(address__location__isnull=False, address__isnull=False)],
                'representatives': UserProfile.objects.filter(role='manager')
            }
            return JsonResponse(final_data)
        return JsonResponse({'success': False, 'error_message': u'Только администратор имеет доступ к странице'})
    return JsonResponse({'success': False, 'error_message': u'Ошибка при просмотре профиля пользователя'})

@login_required(login_url='/sign_in/')
def depot_add(request):
    if hasattr(request.user, 'profile'):
        if request.user.profile.role == 'producer':
            depot_address = Address.objects.create(
                full_address=request.POST['full_address'],
                index=request.POST['index'],
                region=request.POST['region'],
                city=request.POST['city'],
                street=request.POST['street'],
                house=request.POST['house'],
                block=request.POST['block'],
                structure=request.POST['structure'],
                flat=request.POST['flat'],
                location=Point(float(request.POST['lng']), float(request.POST['lat'])),
            )

            ProducerDepot.objects.create(
                producer_id=request.user.id,
                address=depot_address
            )
            return redirect(profile)
        return JsonResponse({'success': False, 'error_message': u'Только производитель может добавлять склады'})
    return JsonResponse({'success': False, 'error_message': u'Ошибка при просмотре профиля пользователя'})

@login_required(login_url='/sign_in/')
def depot_del(request, pk):
    if hasattr(request.user, 'profile'):
        if request.user.profile.role == 'producer':
            try:
                depot = ProducerDepot.objects.get(pk=pk)
            except ProducerDepot.DoesNotExist:
                return JsonResponse({'success': False, 'error_message': u'Склад не существует'})

            if depot.producer_id != request.user.id:
                return JsonResponse({'success': False, 'error_message': u'Склад вам не принадлежит'})

            depot.delete()
            return redirect(profile)
        return JsonResponse({'success': False, 'error_message': u'Только производитель может удалять склады'})
    return JsonResponse({'success': False, 'error_message': u'Ошибка при просмотре профиля пользователя'})

@login_required(login_url='/sign_in/')
def depot_edit(request, pk):
    if hasattr(request.user, 'profile'):
        if request.user.profile.role == 'producer':
            try:
                depot = ProducerDepot.objects.get(pk=pk, producer_id=request.user.id)
                depot.address.full_address = request.POST['full_address']
                depot.address.index = request.POST['index']
                depot.address.region = request.POST['region']
                depot.address.city = request.POST['city']
                depot.address.street = request.POST['street']
                depot.address.house = request.POST['house']
                depot.address.block = request.POST['block']
                depot.address.structure = request.POST['structure']
                depot.address.flat = request.POST['flat']
                depot.address.save()
            except ProducerDepot.DoesNotExist:
                return JsonResponse({'success': False, 'error_message': u'Редактируемый склад не найден'})
            return redirect(profile)
        return JsonResponse({'success': False, 'error_message': u'Только производитель может редактировать склады'})
    return JsonResponse({'success': False, 'error_message': u'Ошибка при просмотре профиля пользователя'})



def product_search(request):
    search_string = request.GET['search_string']
    results = ProductCard.objects.filter(name__search=search_string)
    if 'sorting' in request.GET:
        if int(request.GET['sorting']) == 1:
            results.order_by('pk')
        if int(request.GET['sorting']) == 2 or int(request.GET['sorting']) == 3:
            results.order_by('category_id')
        if int(request.GET['sorting']) == 4:
            results.order_by('price')
    else:
        results.order_by('pk')

    return JsonResponse({
        'products': results,
        'trade_points': TradePoint.objects.filter(customer_id=request.user.id)
    })



@login_required(login_url='/sign_in/')
def product_edit(request, pk):
    """
        todo: only owners should be allowed to edit products
    """
    if hasattr(request.user, 'profile'):
        if request.user.profile.role == 'producer':
            try:
                product = ProductCard.objects.get(pk=pk)
            except ProductCard.DoesNotExist:
                return JsonResponse({'success': False, 'error_message': u'Редактируемый товар не найден'})
            if product.product_depot.producer_id != request.user.id:
                return JsonResponse({'success': False, 'error_message': u'Редактировать можно только свои товары'})
            product.name = request.POST['name']
            product.weight = request.POST['weight']
            product.height = request.POST['height']
            product.width = request.POST['width']
            product.length = request.POST['length']
            product.producer_price = request.POST['producer_price']
            product.customer_price = ProductCard.calculate_customer_price(float(request.POST['producer_price']))
            product.minimum_amount = request.POST['minimum_amount']
            product.expiration_date = request.POST.get('expiration_date', '')
            product.expiration_type_id = request.POST['expiration_type']
            product.description = request.POST['description']

            product.category_id = request.POST['category']
            product.save()

            if 'product_pic' in request.FILES:
                # todo: При изменении файла его было бы неплохо удалять
                product.image.save(str(uuid.uuid4()), request.FILES['product_pic'])
                product.save()
            return redirect(my_products)
        return JsonResponse({'success': False, 'error_message': u'Только производитель может добавлять товар'})
    return JsonResponse({'success': False, 'error_message': u'Ошибка при просмотре профиля пользователя'})


@login_required(login_url='/sign_in/')
def product_del(request, pk):
    if request.user.profile:
        if request.user.profile.role == 'producer':
            try:
                product = ProductCard.objects.get(pk=pk)
            except ProductCard.DoesNotExist:
                return JsonResponse({'error_message': u'Редактируемый товар не найден'})
            if not product.product_depot.producer_id == request.user.id:
                return JsonResponse({'error_message': u'Только производитель может удалять свой товар'})
            product.delete()
            return redirect(my_products)
        return JsonResponse({'success': False, 'error_message': u'Только производитель может добавлять товар'})
    return JsonResponse({'success': False, 'error_message': u'Ошибка при просмотре профиля пользователя'})


@login_required(login_url='/sign_in/')
def product_add(request):
    if hasattr(request.user, 'profile'):
        if request.user.profile.role == 'producer':
            if request.method == 'POST':
                new_product = ProductCard.objects.create(
                    category_id=request.POST['subcategory'],
                    product_depot_id=request.POST['product_depot'],
                    name=request.POST['name'],
                    expiration_date=request.POST.get('expiration_date', ''),
                    expiration_type_id=request.POST['expiration_type'],
                    producer_price=request.POST['producer_price'],
                    customer_price=ProductCard.calculate_customer_price(float(request.POST['producer_price'])),
                    minimum_amount=request.POST['minimum_amount'],
                    weight=request.POST['weight'],
                    barcode=request.POST['barcode'],
                    height=request.POST['height'],
                    width=request.POST['width'],
                    length=request.POST['length'],
                    description=request.POST['description'],
                )

                if 'product_pic' in request.FILES:
                    # todo: При изменении файла его было бы неплохо удалять
                    new_product.image.save(str(uuid.uuid4()), request.FILES['product_pic'])
                    new_product.save()
                return redirect(my_products)

            return JsonResponse({
                'categories': Category.objects.filter(pid__isnull=True),
                'depots': ProducerDepot.objects.filter(producer_id=request.user.id),
                'expiration_values': ExpirationValue.objects.all(),
            })
        return JsonResponse({'success': False, 'error_message': u'Только производитель может добавлять товар'})
    return JsonResponse({'success': False, 'error_message': u'Ошибка при просмотре профиля пользователя'})


@csrf_exempt
@login_required(login_url='/sign_in/')
def my_products_import(request):
    if not request.user.profile.role == 'producer':
        return JsonResponse({'success': False, 'error_msg': u'Только производители могут импортировать продукты'})

    if 'import_file' in request.FILES:
        depot = ProducerDepot.objects.filter(producer=request.user.id).order_by('pk').first()
        if not depot:
            return JsonResponse({'success': False, 'error_msg': u'Заведите хотя-бы один склад'})
        ws = openpyxl.load_workbook(request.FILES['import_file']).get_active_sheet()
        result = {
            'success': True,
            'processed_products': 0,
            'unprocessed_products': 0,
        }
        for excel_row in ws.iter_rows(min_row=6):
            try:
                subcat = Category.objects.get(name=excel_row[5].value)
                ProductCard.objects.create(
                    name=excel_row[1].value,
                    producer_price=excel_row[2].value,
                    customer_price=ProductCard.calculate_customer_price(float(excel_row[2].value)),
                    minimum_amount=excel_row[3].value,
                    category=subcat,
                    expiration_type_id=1,
                    barcode=excel_row[7].value,
                    pack_amount=excel_row[8].value,
                    weight=excel_row[9].value,
                    length=excel_row[10].value,
                    width=excel_row[11].value,
                    height=excel_row[12].value,
                    product_depot=depot,
                    description=excel_row[13].value,
                )
                result['processed_products'] += 1
            except Exception as e:
                result['unprocessed_products'] += 1
        return JsonResponse(result)
    return JsonResponse({'success': False, 'error_msg': u'Не удалось прочесть файл'})


@login_required(login_url='/sign_in/')
def my_products(request):
    if hasattr(request.user, 'profile'):
        if request.user.profile.role == 'producer':
            products = ProductCard.objects.filter(product_depot__producer_id=request.user.id)
            sort_type = request.GET.get('sort', 'default') #' if 'sort' in request.GET else 'default'
            if 'sort' in request.GET:
                if request.GET['sort'] == 'category':
                    products = products.order_by('category__pid__name')
                if request.GET['sort'] == 'subcategory':
                    products = products.order_by('category__name')

            page = request.GET.get('page', 1)
            paginator = Paginator(products, 10)

            try:
                products = paginator.page(page)
            except PageNotAnInteger:
                products = paginator.page(1)
            except EmptyPage:
                products = paginator.page(paginator.num_pages)

            return JsonResponse({
                'products': products,
                'categories': Category.objects.filter(pid__isnull=True),
                'depots': ProducerDepot.objects.filter(producer_id=request.user.id),
                'expiration_values': ExpirationValue.objects.all(),
                'sort_type': sort_type,
                'current_page': page
            })
        return JsonResponse({'success': False, 'error_message': u'Только производитель может просматривать свои товары'})
    return JsonResponse({'success': False, 'error_message': u'Ошибка при просмотре профиля пользователя'})

@login_required(login_url='/sign_in/')
def my_previous_deals(request):
    if hasattr(request.user, 'profile'):
        if request.user.profile.role == 'producer':
            document = openpyxl.load_workbook(filename=os.path.join(settings.DOCS_ROOT, 'orders_export.xlsx'))
            ws = document.get_active_sheet()

            for index, order_unit in enumerate(OrderUnit.objects.filter(producer_id=request.user.id)):
                ws['A' + str(index + 2)] = order_unit.product.barcode
                ws['B' + str(index + 2)] = order_unit.product.name
                ws['C' + str(index + 2)] = order_unit.customer.profile.company_name
                ws['D' + str(index + 2)] = order_unit.order.trade_point.address.castrate_nicely()
                ws['E' + str(index + 2)] = order_unit.order.created.strftime('%d. %m .%Y')
                ws['F' + str(index + 2)] = order_unit.amount
                ws['G' + str(index + 2)] = order_unit.price
                ws['H' + str(index + 2)] = order_unit.calculate_sum()

            document.save(os.path.join(settings.MEDIA_ROOT, 'generated_docs', 'my_orders_{}.xlsx'.format(request.user.id)))
            return redirect('/current_orders')
        return JsonResponse({'success': False,  'error_message': u'Только поставщик производить выгрузку'})
    return JsonResponse({'success': False,  'error_message': u'Ошибка при просмотре профиля пользователя'})

@login_required(login_url='/sign_in/')
def subcategories(request, pk):
    try:
        current_category = Category.objects.get(pk=pk)
        sub_cats = []
        for sub_cat in Category.objects.filter(pid=pk):
            sub_cats.append({
                'id': sub_cat.id,
                'name': sub_cat.name,
                'get_image_url': sub_cat.get_image_url(),
                'unseen': sub_cat.productcard_set.count() - sub_cat.productcard_set.filter(seen__id=request.user.id).count()
            })

        return JsonResponse({
            'category': current_category,
            'subcategories': sub_cats,
        })
    except Category.DoesNotExist:
        return JsonResponse({'success': False, 'error_message': u'Категория не найдена'})

@login_required(login_url='/sign_in/')
def trade_point_add(request):
    if request.user.profile:
        if request.user.profile.role == 'customer':
            depot_address = Address.objects.create(
                full_address=request.POST['full_address'],
                index=request.POST['index'],
                region=request.POST['region'],
                city=request.POST['city'],
                street=request.POST['street'],
                house=request.POST['house'],
                block=request.POST['block'],
                structure=request.POST['structure'],
                flat=request.POST['flat'],
                location=Point(float(request.POST['lng']), float(request.POST['lat'])),
            )

            TradePoint.objects.create(
                customer_id=request.user.id,
                address=depot_address
            )
            return redirect(profile)
        return JsonResponse({'success': False,  'error_message': u'Только заказчик может добавлять торговые точки'})
    return JsonResponse({'success': False,  'error_message': u'Ошибка при просмотре профиля пользователя'})

@login_required(login_url='/sign_in/')
def trade_point_edit(request, pk):
    if hasattr(request.user, 'profile'):
        if request.user.profile.role == 'customer':
            try:
                depot = TradePoint.objects.get(pk=pk, customer_id=request.user.id)
                depot.address.full_address = request.POST['full_address']
                depot.address.index = request.POST['index']
                depot.address.region = request.POST['region']
                depot.address.city = request.POST['city']
                depot.address.street = request.POST['street']
                depot.address.house = request.POST['house']
                depot.address.block = request.POST['block']
                depot.address.structure = request.POST['structure']
                depot.address.flat = request.POST['flat']
                depot.address.save()
            except TradePoint.DoesNotExist:
                return JsonResponse({'success': False,  'error_message': u'Редактируемая торговая точка не найдена'})
            return redirect(profile)
        return JsonResponse({'success': False,  'error_message': u'Только заказчик может редактировать торговые точки'})
    return JsonResponse({'success': False,  'error_message': u'Ошибка при просмотре профиля пользователя'})


@login_required(login_url='/sign_in/')
def trade_point_del(request, pk):
    if hasattr(request.user, 'profile'):
        if request.user.profile.role == 'customer':
            try:
                trade_point = TradePoint.objects.get(pk=pk)
            except TradePoint.DoesNotExist:
                return JsonResponse({'success': False, 'error_message': u'Торговая точка не существует'})

            if trade_point.customer_id != request.user.id:
                return JsonResponse({'success': False, 'error_message': u'Торговая точка вам не пренадлежит'})

            trade_point_orders = Order.objects.filter(trade_point_id=pk).count()
            if trade_point_orders > 0:
                return JsonResponse({'success': False, 'error_message': u'У торговой точки есть заказы'})

            trade_point.delete()

            return redirect(profile)
        return JsonResponse({'success': False, 'error_message': u'Только заказчик может редактировать торговые точки'})
    return JsonResponse({'success': False, 'error_message': u'Ошибка при просмотре профиля пользователя'})


@login_required(login_url='/sign_in/')
def basket(request):
    if hasattr(request.user, 'profile'):
        if request.user.profile.role == 'customer' or request.user.profile.role == 'manager':
            order_units = OrderUnit.objects.filter(order__isnull=True, customer_id=request.user.id).order_by('pk')
            """
            if request.user.profile.role == 'customer':
                trade_points = TradePoint.objects.filter(customer_id=request.user.id).order_by('pk')
            if request.user.profile.role == 'manager':
                trade_points = TradePoint.objects.filter(territory__representative_id=request.user.id).order_by('pk')
            """
            return JsonResponse({
                'order_units': order_units,
                #'trade_points': trade_points
            })
        return JsonResponse({'success': False, 'error_message': u'Только заказчики и торговые представители могут просматривать корзину'})
    return JsonResponse({'success': False, 'error_message': u'Недостаточно прав для совершения операции'})

def basket_trade_points(request):
    if hasattr(request.user, 'profile'):
        if request.user.profile.role == 'customer' or request.user.profile.role == 'manager':
            if request.user.profile.role == 'customer':
                trade_points = TradePoint.objects.filter(customer_id=request.user.id).order_by('pk')
            if request.user.profile.role == 'manager':
                trade_points = TradePoint.objects.filter(territory__representative_id=request.user.id).order_by('pk')

            if 'term' in request.GET:
                trade_points = trade_points.filter(address__full_address__icontains=request.GET['term'])
            return JsonResponse([{'id': tp.id, 'name': tp.customer_and_address(), 's_name': tp.address.full_address}
                                 for tp in trade_points], safe=False)
        return JsonResponse({'success': False, 'error_message': u'Только заказчики и торговые представители могут просматривать корзину'})
    return JsonResponse({'success': False, 'error_message': u'Недостаточно прав для совершения операции'})


@login_required(login_url='/sign_in/')
def order_unit_add(request):
    if hasattr(request.user, 'profile'):
        if request.user.profile.role == 'customer' or request.user.profile.role == 'manager':
            try:
                product = ProductCard.objects.get(pk=request.POST['product'])
                endpoint = '/products/{}/'.format(product.category_id)
            except ProductCard.DoesNotExist:
                return JsonResponse({'success': False, 'error_message': u'Не найден добавляемый продукт'})


            """Если в корзине уже есть товар - меняем цену"""
            if OrderUnit.objects.filter(order__isnull=True, product_id=product.id, customer_id=request.user.id).count() > 0:
                order_unit = OrderUnit.objects.get(order__isnull=True, product_id=product.id, customer_id=request.user.id)
                order_unit.amount = int(request.POST['amount'])
                order_unit.save()
                return redirect(endpoint)

            order_unit = OrderUnit.objects.create(
                producer_id=product.product_depot.producer.id,
                customer_id=request.user.id,
                product=product,
                amount=int(request.POST['amount']),
                price=product.producer_price,
            )
            order_unit.save()
            return redirect(endpoint)
        return JsonResponse({'success': False, 'error_message': u'Только заказчики или торговые представители могут добавлять позиции заказа'})
    return JsonResponse({'success': False, 'error_message': u'Недостаточно прав для совершения операции'})


@login_required(login_url='/sign_in/')
def order_unit_edit(request, pk):
    if hasattr(request.user, 'profile'):
        if request.user.profile.role == 'customer' or request.user.profile.role == 'manager':
            try:
                order_unit = OrderUnit.objects.get(pk=pk)
                if order_unit.customer_id != request.user.id:
                    return JsonResponse({'success': False,'error_message': u'Редактировать можно только свои позиции заказа'})
                order_unit.amount = request.POST['amount']
                order_unit.remark = request.POST['remark']
                order_unit.price = order_unit.product.producer_price
                order_unit.save()
            except OrderUnit.DoesNotExist:
                return JsonResponse({'success': False, 'error_message': u'Позиция заказа не существует'})

            return redirect(basket)
        return JsonResponse({'success': False,'error_message': u'Только заказчики и торговые представители могут редактировать позиции заказа'})
    return JsonResponse({'success': False,'error_message': u'Недостаточно прав для совершения операции'})



@login_required(login_url='/sign_in/')
def order_unit_del(request, pk):
    if hasattr(request.user, 'profile'):
        if request.user.profile.role == 'customer' or request.user.profile.role == 'manager':
            try:
                order_unit = OrderUnit.objects.get(pk=pk)
                if order_unit.customer_id != request.user.id:
                    return JsonResponse({'success': False, 'error_message': u'Удалять можно только свои позиции заказа'})
                order_unit.delete()
            except OrderUnit.DoesNotExist:
                return JsonResponse({'success': False, 'error_message': u'Позиция заказа не существует'})

            return redirect(basket)
        return JsonResponse({'success': False, 'error_message': u'Только заказчики или торговые представители могут удалять позиции заказа'})
    return JsonResponse({'success': False, 'error_message': u'Недостаточно прав для совершения операции'})


@login_required(login_url='/sign_in/')
def perform_order(request):
    if hasattr(request.user, 'profile'):
        if request.user.profile.role == 'customer' or request.user.profile.role == 'manager':
            try:
                tp = TradePoint.objects.get(pk=request.POST['trade_point'])
            except TradePoint.DoesNotExist:
                return JsonResponse({'success': False, 'error_message': u'Не найдена торговая точка'})

            cost_of_basket = 0
            for order_unit in OrderUnit.objects.filter(order__isnull=True, customer_id=request.user.id):
                cost_of_basket += order_unit.amount * order_unit.price

            order_payment = OrderPayment.objects.create(price=cost_of_basket, customer_id=tp.customer_id)

            # todo: Тут можно покрасивее переписать, group_by
            order_unit_producer_id = 0
            for order_unit in (OrderUnit.objects.filter(order__isnull=True, customer_id=request.user.id).order_by('producer_id')):
                if order_unit.producer_id != order_unit_producer_id:
                    order = Order.objects.create(
                        customer_id=tp.customer_id,
                        trade_point_id=tp.id,
                        producer_id=order_unit.producer_id,
                        payment=order_payment)
                    order_unit.order_id = order.id
                    order_unit.save()
                else:
                    order = Order.objects.filter(
                        customer_id=tp.customer_id,
                        trade_point_id=tp.id,
                        producer_id=order_unit.producer_id).order_by('id').reverse()[0]
                    order_unit.order_id = order.id
                    order_unit.save()
                order_unit_producer_id = order_unit.producer_id

            if request.user.profile.role == 'customer':
                return redirect('/payment_type/{}/'.format(order_payment.id))
            return redirect('/my_clients')
        return JsonResponse({'success': False, 'error_message': u'Только заказчик может просматривать свои заказы'})
    return JsonResponse({'success': False, 'error_message': u'Недостаточно прав для совершения операции'})



@csrf_exempt
@login_required(login_url='/sign_in')
def current_orders(request):
    if not hasattr(request.user, 'profile'):
        return JsonResponse({'success': False, 'error_msg': u'Недостаточно прав для совершения операции'})
    if not request.user.profile.role == 'producer':
        return JsonResponse({'success': False, 'error_msg': u'Только поставщики могут просматривать заказы на карте'})
    return JsonResponse({cur_order.id: {'name': cur_order.trade_point.address.castrate_nicely(),
                                        'location': {
                                            'lat': cur_order.trade_point.address.location.y,
                                            'lng': cur_order.trade_point.address.location.x}
                                        } for cur_order in
                         Order.objects.filter(producer_id=request.user.id, order_status__in=(1, 2, 4, 6, 8),
                                              trade_point__address__location__isnull=False).order_by('created')})


@login_required(login_url='/sign_in/')
def order_history(request):
    if hasattr(request.user, 'profile'):
        cur_orders = None
        if request.user.profile.role == 'customer':
            cur_orders = Order.objects.filter(customer_id=request.user.id, order_status__in=(3, 5, 7))

        if request.user.profile.role == 'producer':
            cur_orders = Order.objects.filter(producer_id=request.user.id, order_status__in=(3, 5, 7))
        return JsonResponse({
            'current_orders': cur_orders
        })
    return JsonResponse({'success': False, 'error_message': u'Недостаточно прав для совершения операции'})

@login_required(login_url='/sign_in/')
def order(request, pk):
    try:
        current_order = Order.objects.get(pk=pk)
        if not (request.user.id == current_order.producer_id or request.user.id == current_order.customer_id):
            return JsonResponse({'success': False, 'error_message': u'Просматривать можно только свои заказы'})

        order_units = OrderUnit.objects.filter(order=current_order)

    except Order.DoesNotExist:
        return JsonResponse({'success': False, 'error_message': u'Заказ не существует'})
    return JsonResponse({'order': current_order, 'order_units': order_units})

@login_required(login_url='/sign_in/')
def set_status_sent(request, pk):
    if not request.user.profile.role == 'producer':
        return JsonResponse({'success': False, 'error_message': u'Только поставщик может подтверждать отправку заказа'})
    try:
        current_order = Order.objects.get(pk=pk)
        if not request.user.id == current_order.producer_id:
            return JsonResponse({'success': False, 'error_message': u'Только поставщик может подтверждать отправку заказа'})
        current_order.set_status_sent()
        return redirect(order, pk)
    except Order.DoesNotExist:
        return JsonResponse({'success': False, 'error_message': u'Заказ не существует'})



@login_required(login_url='/sign_in/')
def set_status_delivered(request, pk):
    if not request.user.profile.role == 'customer':
        return JsonResponse({'success': False, 'error_message': u'Только заказчик может подтверждать доставку заказа'})
    try:
        current_order = Order.objects.get(pk=pk)
        if not request.user.id == current_order.customer_id:
            return JsonResponse({'success': False, 'error_message': u'Только заказчик может подтверждать доставку заказа'})
        current_order.set_status_delivered()
        return redirect(order, pk)
    except Order.DoesNotExist:
        return JsonResponse({'success': False, 'error_message': u'Заказ не существует'})



@login_required(login_url='/sign_in/')
def order_payment(request, pk):
    if not hasattr(request.user, 'profile'):
        return JsonResponse({'success': False, 'error_message': u'Не создан профиль пользователя'})
    if request.user.profile.role != 'customer':
        return JsonResponse({'success': False, 'error_message': u'Только заказчики могут просматривать страницу оплаты заказов'})
    try:
        payment = OrderPayment.objects.get(pk=pk)
        order = Order.objects.filter(payment__id=payment.id).first()
    except Order.DoesNotExist:
        return JsonResponse({'success': False, 'error_message': u'Заказ не существует'})

    document = Document(os.path.join(settings.DOCS_ROOT, 'invoice_payment.docx'))
    if not request.user.id == payment.customer_id:
        return JsonResponse({'success': False, 'error_message': u'Просматривать можно только свои заказы'})

    replacements = {
        '{{sum}}': payment.price,
        '{{order_number}}': payment.id,
        '{{order_date}}': payment.created,
        '{{company}}': request.user.profile.company_name,
        '{{inn}}': request.user.profile.inn,
        '{{kpp}}': request.user.profile.kpp,
        '{{address}}': order.trade_point.address.castrate_nicely()
    }

    for table_index, table in enumerate(document.tables):
        for row_index, row in enumerate(table.rows):
            print row_index
            for cell in row.cells:
                for rep_ind, rep_val in replacements.iteritems():
                    cell.text = cell.text.replace(rep_ind, unicode(rep_val))

    # TODO: Доступ к документам ограничить бы
    document.save(os.path.join(settings.MEDIA_ROOT, 'generated_docs', 'invoice_payment_{}.docx'.format(payment.id)))

    return JsonResponse({'payment': payment})

@login_required(login_url='/sign_in/')
def direct_payment(request, pk):
    try:
        payment = OrderPayment.objects.get(pk=pk)
    except OrderPayment.DoesNotExist:
        return JsonResponse({'success': False, 'error_message': u'Платёж не найден'})
    if not request.user.id == payment.customer_id:
        return JsonResponse({'success': False, 'error_message': u'Только создатель заказа может его оплачивать'})

    if 'document' not in request.FILES:
        return JsonResponse({'success': False, 'error_message': u'Прикрепите платёжное поручение'})

    DirectPayment.objects.create(
        payment=payment,
        document=request.FILES['document'],
        document_id=request.POST['document_id'],
        date=datetime.datetime.strptime(request.POST['date'], "%d.%m.%Y").date()
    )

    return redirect(current_orders)


@login_required(login_url='/sign_in/')
def bank_payment(request, pk):
    try:
        payment = OrderPayment.objects.get(pk=pk)
    except OrderPayment.DoesNotExist:
        return JsonResponse({'success': False, 'error_message': u'Платёж не найден'})
    if not request.user.id == payment.customer_id:
        return JsonResponse({'success': False, 'error_message': u'Только создатель заказа может его оплачивать'})

    return JsonResponse({'payment': payment})


@login_required(login_url='/sign_in/')
def payment_type(request, pk):
    try:
        op = OrderPayment.objects.get(pk=pk)
    except OrderPayment.DoesNotExist:
        return JsonResponse({'success': False, 'error_message': u'Не найден счёт на оплату'})
    return JsonResponse({'order_payment': op})


@login_required(login_url='/sign_in/')
def payment_type_pick(request):
    try:
        op = OrderPayment.objects.get(pk=request.POST['order_payment'])
    except OrderPayment.DoesNotExist:
        return JsonResponse({'success': False, 'error_message': u'Не найден счёт на оплату'})
    if payment_type == 0:
        return redirect('/order_payment/{}/'.format(op.id))
    return redirect('/bank_payment/{}/'.format(op.id))

    #return render(request, 'payment_type.html', {'order_payment': op})



@csrf_exempt
def payment_notification(request):
    PaymentNotification.objects.create(response=json.dumps(request.body))
    return redirect('/')


@csrf_exempt
def success_redirect(request):
    Success.objects.create(response=json.dumps(request.body))
    return redirect('/')


@csrf_exempt
def failure_redirect(request):
    Success.objects.create(response=json.dumps(request.body))     #Failure???
    return redirect('/')



"""
def apply_context(result_dict, user):
    if request.user.is_authenticated and not request.user.is_superuser:
        if hasattr(request.user, 'profile'):
            if request.user.profile.role == 'customer' or request.user.profile.role == 'manager':
                cost_of_basket = 0
                for order_unit in OrderUnit.objects.filter(order__isnull=True, customer_id=request.user.id):
                    cost_of_basket += order_unit.amount * order_unit.price
                return {
                    'basket_items': OrderUnit.objects.filter(order__isnull=True, customer_id=request.user.id).count(),
                    'basket_price': cost_of_basket
                }
    return {'basket_items': 0, 'basket_price': 0}
"""
sign_in = SignIn.as_view()
