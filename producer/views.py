# -*- coding: utf-8 -*-
from __future__ import unicode_literals

import os
import uuid
import openpyxl

from django.conf import settings
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.contrib.gis.geos import Point

from core.views import profile
from core.models import Address
from catalogs.models import Category, ExpirationValue
from producer.models import ProductCard, ProducerDepot
from customer.models import Order, OrderUnit

"""
CELL_LETTER = {
    0: 'A',
    1: 'B',
    2: 'C',
    3: 'D',
    4: 'E',
    5: 'F',
    6: 'G',
    7: 'H',
    8: 'I',
    9: 'J'
}
"""

@login_required(login_url='/sign_in/')
def my_products(request):
    if request.user.profile:
        if request.user.profile.role == 'producer':
            return render(request, 'my_products.html', {
                'products': ProductCard.objects.filter(product_depot__producer_id=request.user.id).order_by('pk'),
                'categories': Category.objects.filter(pid__isnull=True),
                'depots': ProducerDepot.objects.filter(producer_id=request.user.id),
                'expiration_values': ExpirationValue.objects.all(),
            })
        return render(request, '500.html', {'error_message': u'Только производитель может просматривать свои товары'})
    return render(request, '500.html', {'error_message': u'Ошибка при просмотре профиля пользователя'})


@csrf_exempt
@login_required(login_url='/sign_in/')
def my_products_import(request):
    if not request.user.profile.role == 'producer':
        return JsonResponse({'success': False, 'error_msg': u'Только производители могут импортировать продукты'})

    if 'import_file' in request.FILES:
        depot = ProducerDepot.objects.filter(producer=request.user.id).order_by('pk').first()
        if not depot:
            return JsonResponse({'success': False, 'error_msg': u'Заведите хотя-бы один склад'})
        ws = openpyxl.load_workbook(request.FILES['import_file']).get_active_sheet()
        result = {
            'success': True,
            'processed_products': 0,
            'unprocessed_products': 0,
        }
        for excel_row in ws.iter_rows(min_row=6):
            try:
                subcat = Category.objects.get(name=excel_row[5].value)
                ProductCard.objects.create(
                    name=excel_row[1].value,
                    producer_price=excel_row[2].value,
                    customer_price=ProductCard.calculate_customer_price(float(excel_row[2].value)),
                    minimum_amount=excel_row[3].value,
                    category=subcat,
                    expiration_type_id=1,
                    barcode=excel_row[7].value,
                    pack_amount=excel_row[8].value,
                    weight=excel_row[9].value,
                    length=excel_row[10].value,
                    width=excel_row[11].value,
                    height=excel_row[12].value,
                    product_depot=depot,
                    description=excel_row[13].value,
                )
                result['processed_products'] += 1
            except Exception as e:
                result['unprocessed_products'] += 1
        return JsonResponse(result)
    return JsonResponse({'success': False, 'error_msg': u'Не удалось прочесть файл'})


@login_required(login_url='/sign_in/')
def product_add(request):
    if request.user.profile:
        if request.user.profile.role == 'producer':
            if request.method == 'POST':
                new_product = ProductCard.objects.create(
                    category_id=request.POST['subcategory'],
                    product_depot_id=request.POST['product_depot'],
                    name=request.POST['name'],
                    expiration_date=request.POST.get('expiration_date', ''),
                    expiration_type_id=request.POST['expiration_type'],
                    producer_price=request.POST['producer_price'],
                    customer_price=ProductCard.calculate_customer_price(float(request.POST['producer_price'])),
                    minimum_amount=request.POST['minimum_amount'],
                    weight=request.POST['weight'],
                    barcode=request.POST['barcode'],
                    height=request.POST['height'],
                    width=request.POST['width'],
                    length=request.POST['length'],
                    description=request.POST['description'],
                )

                if 'product_pic' in request.FILES:
                    # todo: При изменении файла его было бы неплохо удалять
                    new_product.image.save(str(uuid.uuid4()), request.FILES['product_pic'])
                    new_product.save()
                return redirect(my_products)

            return render(request, 'product_add.html', {
                'categories': Category.objects.filter(pid__isnull=True),
                'depots': ProducerDepot.objects.filter(producer_id=request.user.id),
                'expiration_values': ExpirationValue.objects.all(),
            })
        return render(request, '500.html', {'error_message': u'Только производитель может добавлять товар'})
    return render(request, '500.html', {'error_message': u'Ошибка при просмотре профиля пользователя'})


@login_required(login_url='/sign_in/')
def product_edit(request, pk):
    """
        todo: only owners should be allowed to edit products
    """
    if request.user.profile:
        if request.user.profile.role == 'producer':
            try:
                product = ProductCard.objects.get(pk=pk)
            except ProductCard.DoesNotExist:
                return render(request, '500.html', {'error_message': u'Редактируемый товар не найден'})
            if product.product_depot.producer_id != request.user.id:
                return render(request, '500.html', {'error_message': u'Редактировать можно только свои товары'})
            product.name = request.POST['name']
            product.weight = request.POST['weight']
            product.height = request.POST['height']
            product.width = request.POST['width']
            product.length = request.POST['length']
            product.producer_price = request.POST['producer_price']
            product.customer_price = ProductCard.calculate_customer_price(float(request.POST['producer_price']))
            product.minimum_amount = request.POST['minimum_amount']
            product.expiration_date = request.POST.get('expiration_date', '')
            product.expiration_type_id = request.POST['expiration_type']
            product.description = request.POST['description']

            product.category_id = request.POST['category']
            product.save()

            if 'product_pic' in request.FILES:
                # todo: При изменении файла его было бы неплохо удалять
                product.image.save(str(uuid.uuid4()), request.FILES['product_pic'])
                product.save()
            return redirect(my_products)
        return render(request, '500.html', {'error_message': u'Только производитель может добавлять товар'})
    return render(request, '500.html', {'error_message': u'Ошибка при просмотре профиля пользователя'})


@login_required(login_url='/sign_in/')
def product_del(request, pk):
    if request.user.profile:
        if request.user.profile.role == 'producer':
            try:
                product = ProductCard.objects.get(pk=pk)
            except ProductCard.DoesNotExist:
                return render(request, '500.html', {'error_message': u'Редактируемый товар не найден'})
            if not product.product_depot.producer_id == request.user.id:
                return render(request, '500.html', {'error_message': u'Только производитель может удалять свой товар'})
            product.delete()
            return redirect(my_products)
        return render(request, '500.html', {'error_message': u'Только производитель может добавлять товар'})
    return render(request, '500.html', {'error_message': u'Ошибка при просмотре профиля пользователя'})


@login_required(login_url='/sign_in/')
def depot_add(request):
    if request.user.profile:
        if request.user.profile.role == 'producer':
            depot_address = Address.objects.create(
                full_address=request.POST['full_address'],
                index=request.POST['index'],
                region=request.POST['region'],
                city=request.POST['city'],
                street=request.POST['street'],
                house=request.POST['house'],
                block=request.POST['block'],
                structure=request.POST['structure'],
                flat=request.POST['flat'],
                location=Point(float(request.POST['lng']), float(request.POST['lat'])),
            )

            ProducerDepot.objects.create(
                producer_id=request.user.id,
                address=depot_address
            )
            return redirect(profile)
        return render(request, '500.html', {'error_message': u'Только производитель может добавлять склады'})
    return render(request, '500.html', {'error_message': u'Ошибка при просмотре профиля пользователя'})


@login_required(login_url='/sign_in/')
def depot_edit(request, pk):
    if request.user.profile:
        if request.user.profile.role == 'producer':
            try:
                depot = ProducerDepot.objects.get(pk=pk, producer_id=request.user.id)
                depot.address.full_address = request.POST['full_address']
                depot.address.index = request.POST['index']
                depot.address.region = request.POST['region']
                depot.address.city = request.POST['city']
                depot.address.street = request.POST['street']
                depot.address.house = request.POST['house']
                depot.address.block = request.POST['block']
                depot.address.structure = request.POST['structure']
                depot.address.flat = request.POST['flat']
                depot.address.save()
            except ProducerDepot.DoesNotExist:
                return render(request, '500.html', {'error_message': u'Редактируемый склад не найден'})
            return redirect(profile)
        return render(request, '500.html', {'error_message': u'Только производитель может редактировать склады'})
    return render(request, '500.html', {'error_message': u'Ошибка при просмотре профиля пользователя'})


@login_required(login_url='/sign_in/')
def my_previous_deals(request):
    if request.user.profile:
        if request.user.profile.role == 'producer':
            document = openpyxl.load_workbook(filename=os.path.join(settings.DOCS_ROOT, 'orders_export.xlsx'))
            ws = document.get_active_sheet()

            for index, order_unit in enumerate(OrderUnit.objects.filter(producer_id=request.user.id)):
                ws['A' + str(index + 2)] = order_unit.product.barcode
                ws['B' + str(index + 2)] = order_unit.product.name
                ws['C' + str(index + 2)] = order_unit.customer.profile.company_name
                ws['D' + str(index + 2)] = order_unit.order.trade_point.address.castrate_nicely()
                ws['E' + str(index + 2)] = order_unit.order.created.strftime('%d. %m .%Y')
                ws['F' + str(index + 2)] = order_unit.amount
                ws['G' + str(index + 2)] = order_unit.price
                ws['H' + str(index + 2)] = order_unit.calculate_sum()

            document.save(os.path.join(settings.MEDIA_ROOT, 'generated_docs', 'my_orders_{}.xlsx'.format(request.user.id)))
            return redirect('/current_orders')
        return render(request, '500.html', {'error_message': u'Только поставщик производить выгрузку'})
    return render(request, '500.html', {'error_message': u'Ошибка при просмотре профиля пользователя'})