# -*- coding: utf-8 -*-
import os
import openpyxl

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.conf import settings
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required

from customer.models import Order, OrderUnit


@login_required(login_url='/sign_in/')
def current_orders(request):
    if hasattr(request.user, 'profile'):
        cur_orders = None
        document = None
        if request.user.profile.role == 'customer':
            cur_orders = Order.objects.filter(customer_id=request.user.id, order_status__in=(1, 2, 4, 6, 8))

        if request.user.profile.role == 'producer':
            cur_orders = Order.objects.filter(producer_id=request.user.id, order_status__in=(1, 2, 4, 6, 8))
            document = openpyxl.load_workbook(filename=os.path.join(settings.DOCS_ROOT, 'orders_export.xlsx'))
            ws = document.get_active_sheet()

            for index, order_unit in enumerate(OrderUnit.objects.filter(producer_id=request.user.id,
                                                                        order__trade_point__isnull=False,
                                                                        order__trade_point__territory__isnull=False)):
                ws['A' + str(index + 2)] = int(order_unit.product.barcode)
                ws['B' + str(index + 2)] = order_unit.product.name
                ws['C' + str(index + 2)] = order_unit.customer.profile.company_name
                ws['D' + str(index + 2)] = order_unit.order.trade_point.address.castrate_nicely()
                ws['E' + str(index + 2)] = order_unit.order.created.strftime('%d.%m.%Y')
                ws['F' + str(index + 2)] = order_unit.amount
                ws['G' + str(index + 2)] = order_unit.price
                ws['H' + str(index + 2)] = order_unit.calculate_sum()
                ws['I' + str(index + 2)] = order_unit.order.trade_point.territory.name

            document.save(
                os.path.join(settings.MEDIA_ROOT, 'generated_docs', 'my_orders_{}.xlsx'.format(request.user.id)))

        final_data = {'current_orders': cur_orders}
        if document is not None:
            final_data['link_to_excel'] = os.path.join('media', 'generated_docs',
                                                       'my_orders_{}.xlsx'.format(request.user.id))
        return render(request, 'current_orders.html', final_data)
    return render(request, '500.html', {'error_message': u'Недостаточно прав для совершения операции'})


@login_required(login_url='/sign_in/')
def order_history(request):
    if hasattr(request.user, 'profile'):
        cur_orders = None
        if request.user.profile.role == 'customer':
            cur_orders = Order.objects.filter(customer_id=request.user.id, order_status__in=(3, 5, 7))

        if request.user.profile.role == 'producer':
            cur_orders = Order.objects.filter(producer_id=request.user.id, order_status__in=(3, 5, 7))
        return render(request, 'order_history.html', {
            'current_orders': cur_orders
        })
    return render(request, '500.html', {'error_message': u'Недостаточно прав для совершения операции'})


@login_required(login_url='/sign_in/')
def order(request, pk):
    try:
        current_order = Order.objects.get(pk=pk)
        if not (request.user.id == current_order.producer_id or request.user.id == current_order.customer_id):
            return render(request, '500.html', {'error_message': u'Просматривать можно только свои заказы'})

        order_units = OrderUnit.objects.filter(order=current_order)

    except Order.DoesNotExist:
        return render(request, '500.html', {'error_message': u'Заказ не существует'})
    return render(request, 'order.html', {'order': current_order, 'order_units': order_units})


@login_required(login_url='/sign_in/')
def set_status_sent(request, pk):
    if not request.user.profile.role == 'producer':
        return render(request, '500.html', {'error_message': u'Только поставщик может подтверждать отправку заказа'})
    try:
        current_order = Order.objects.get(pk=pk)
        if not request.user.id == current_order.producer_id:
            return render(request, '500.html',
                          {'error_message': u'Только поставщик может подтверждать отправку заказа'})
        current_order.set_status_sent()
        return redirect(order, pk)
    except Order.DoesNotExist:
        return render(request, '500.html', {'error_message': u'Заказ не существует'})


@login_required(login_url='/sign_in/')
def set_status_delivered(request, pk):
    if not request.user.profile.role == 'customer':
        return render(request, '500.html', {'error_message': u'Только заказчик может подтверждать доставку заказа'})
    try:
        current_order = Order.objects.get(pk=pk)
        if not request.user.id == current_order.customer_id:
            return render(request, '500.html', {'error_message': u'Только заказчик может подтверждать доставку заказа'})
        current_order.set_status_delivered()
        return redirect(order, pk)
    except Order.DoesNotExist:
        return render(request, '500.html', {'error_message': u'Заказ не существует'})


@csrf_exempt
@login_required(login_url='/sign_in')
def current_orders_json(request):
    if not hasattr(request.user, 'profile'):
        return JsonResponse({'success': False, 'error_msg': u'Недостаточно прав для совершения операции'})
    if not request.user.profile.role == 'producer':
        return JsonResponse({'success': False, 'error_msg': u'Только поставщики могут просматривать заказы на карте'})
    return JsonResponse({cur_order.id: {'name': cur_order.trade_point.address.castrate_nicely(),
                                        'location': {
                                            'lat': cur_order.trade_point.address.location.y,
                                            'lng': cur_order.trade_point.address.location.x}
                                        } for cur_order in
                         Order.objects.filter(producer_id=request.user.id, order_status__in=(1, 2, 4, 6, 8),
                                              trade_point__address__location__isnull=False)})
